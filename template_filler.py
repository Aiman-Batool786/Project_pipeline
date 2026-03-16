"""
template_filler.py
──────────────────
Fills the Octopia .xlsm template:
  ROW 1  → OCTOPIA | Catégorie | <category_id> | <category_name>
  ROW 11 → first product data row  (subsequent calls append below)

The _build_field_map() reads ROW 5 (FIELD_KEY_ROW) to discover column
positions.  Keys are stored in lowercase-stripped form so lookups are
case-insensitive and whitespace-tolerant.
"""

import json
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

FIELD_KEY_ROW = 5
CATEGORY_ROW  = 1
DATA_ROW      = 11


class TemplateFiller:

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb            = None
        self.ws            = None
        # Maps  lowercase(header_text) → column index (1-based)
        self._field_col: dict[str, int] = {}

    # ── LOAD ──────────────────────────────────────────────────────────────────

    def load_template(self) -> bool:
        try:
            self.wb = load_workbook(self.template_path, keep_vba=True)
            self.ws = self.wb.active
            self._build_field_map()
            print(f"[template] ✅ Loaded  : {self.template_path}")
            print(f"[template] 📋 Sheet   : {self.ws.title}")
            print(f"[template] 📋 Columns : {len(self._field_col)} mapped from ROW {FIELD_KEY_ROW}")
            return True
        except Exception as e:
            print(f"[template] ❌ Load error: {e}")
            return False

    def _build_field_map(self):
        """
        Read ROW 5 and build  lowercase(header) → column index.
        This makes all lookups case/whitespace-insensitive.
        """
        self._field_col = {}
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=FIELD_KEY_ROW, column=col).value
            if val:
                key = str(val).strip().lower()
                self._field_col[key] = col

    def _lookup_col(self, field_key: str) -> int | None:
        """
        Return column index for field_key (case-insensitive).
        Returns None if not found.
        """
        return self._field_col.get(field_key.strip().lower())

    # ── CATEGORY ROW ──────────────────────────────────────────────────────────

    def fill_category_row(self, category_id: str, category_name: str) -> bool:
        """Write OCTOPIA category info to ROW 1."""
        try:
            cat_id   = str(category_id).strip()  if category_id   else "0"
            cat_name = str(category_name).strip() if category_name else "Uncategorized"

            self.ws.cell(row=CATEGORY_ROW, column=1).value = "OCTOPIA"
            self.ws.cell(row=CATEGORY_ROW, column=2).value = "Catégorie"
            self.ws.cell(row=CATEGORY_ROW, column=3).value = cat_id
            self.ws.cell(row=CATEGORY_ROW, column=4).value = cat_name

            print(f"[template] 🏷️  Category ROW {CATEGORY_ROW}: "
                  f"id={cat_id} | name={cat_name}")
            return True
        except Exception as e:
            print(f"[template] ❌ fill_category_row error: {e}")
            return False

    # ── DATA ROW ──────────────────────────────────────────────────────────────

    def _find_next_data_row(self) -> int:
        """Find the first empty row at or below DATA_ROW (checks column C)."""
        for r in range(DATA_ROW, self.ws.max_row + 2):
            if self.ws.cell(row=r, column=3).value is None:
                return r
        return self.ws.max_row + 1

    def fill_product_data(self, mapped_data: dict) -> bool:
        """
        Write mapped_data to the next available data row.

        mapped_data keys must match the column headers in ROW 5
        (comparison is case-insensitive / whitespace-tolerant).
        """
        try:
            row    = self._find_next_data_row()
            filled = 0
            missed = []

            print(f"\n[template] 📝 Writing {len(mapped_data)} fields → ROW {row}...")

            for field_key, value in mapped_data.items():

                col = self._lookup_col(field_key)
                if col is None:
                    missed.append(field_key)
                    continue

                # Normalise value
                if isinstance(value, list):
                    cell_val = " | ".join(str(v) for v in value if v)
                elif isinstance(value, dict):
                    cell_val = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    continue
                else:
                    cell_val = str(value).strip()

                if not cell_val:
                    continue

                self.ws.cell(row=row, column=col).value = cell_val
                filled += 1

            print(f"[template] ✅ {filled} fields written to ROW {row}")

            if missed:
                print(f"[template] ⚠️  {len(missed)} keys not found in template "
                      f"ROW {FIELD_KEY_ROW}: {missed[:10]}"
                      + (" …" if len(missed) > 10 else ""))

            return True

        except Exception as e:
            print(f"[template] ❌ fill_product_data error: {e}")
            return False

    # ── SAVE / CLOSE ──────────────────────────────────────────────────────────

    def save_template(self, output_path: str) -> bool:
        try:
            self.wb.save(output_path)
            print(f"[template] ✅ Saved: {output_path}")
            return True
        except Exception as e:
            print(f"[template] ❌ Save error: {e}")
            return False

    def close(self):
        if self.wb:
            self.wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# CONVENIENCE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def fill_template_for_product(
    template_path: str,
    mapped_data: dict,
    product_id: int,
    output_dir: str = "./filled_templates",
    category_id: str = "0",
    category_name: str = "Uncategorized"
) -> str | None:
    """
    Copy the master template, fill it for one product, and return the output path.

    Args:
        template_path  : path to the master .xlsm file
        mapped_data    : {column_header: value} from data_mapper
        product_id     : used in the output filename
        output_dir     : directory to save the filled file
        category_id    : written to ROW 1, column C
        category_name  : written to ROW 1, column D

    Returns:
        output_path on success, None on failure
    """
    if not os.path.exists(template_path):
        print(f"[template] ❌ Template not found: {template_path}")
        return None

    os.makedirs(output_dir, exist_ok=True)

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename    = f"product_{product_id}_{ts}.xlsm"
    output_path = os.path.join(output_dir, filename)

    # Work on a copy — never modify the master
    shutil.copy2(template_path, output_path)

    filler = TemplateFiller(output_path)

    def _cleanup():
        filler.close()
        if os.path.exists(output_path):
            os.remove(output_path)

    if not filler.load_template():
        _cleanup()
        return None

    if not filler.fill_category_row(category_id, category_name):
        _cleanup()
        return None

    if not filler.fill_product_data(mapped_data):
        _cleanup()
        return None

    if not filler.save_template(output_path):
        filler.close()
        return None

    filler.close()
    return output_path
