"""
data_mapper.py
──────────────
Maps scraped and LLM-enhanced product data into the Octopia xlsm template.

Template column layout (1-indexed):
  Col 1  – GTIN
  Col 2  – Référence vendeur (sellerProductReference)
  Col 3  – Titre* (title)
  Col 4  – Description* (raw scraped text)
  Col 5  – URL image 1*
  Col 6  – Groupe de variation
  Col 7  – Référence de Regroupement des Variants*
  Col 8  – Marque (brand)
  Col 9  – Description marketing (richMarketingDescription → HTML)
  Col 10 – URL image 2
  Col 11 – URL image 3
  Col 12 – URL image 4
  Col 13 – URL image 5
  Col 14 – URL image 6

Row 1:  OCTOPIA | Catégorie | category_code | leaf_category
Row 4:  Human-readable column headers
Row 5:  Field key names
Rows 9+: Product data rows
"""

import os
import re
import shutil
from datetime import datetime

import openpyxl


# ─────────────────────────────────────────
# CATEGORY PARSING
# ─────────────────────────────────────────

def parse_category(category_name: str) -> tuple[str, str]:
    """
    Given a full category string like:
        "ADULT - EROTIC/ARTICLES WITH SEXUAL CONNOTATIONS/DISHWASHER"
    Return (category_code, leaf_category).

    If category_name already contains a code prefix (e.g. "0V0702 | DISHWASHER"),
    split on | and return both parts.
    Otherwise return ("", category_name.split("/")[-1].strip())
    """
    if not category_name:
        return ("", "")

    # Already formatted as "code | leaf"
    if "|" in category_name:
        parts = [p.strip() for p in category_name.split("|")]
        return (parts[0], parts[-1])

    # Raw CSV format: extract leaf (last segment after /)
    leaf = category_name.split("/")[-1].strip()
    return ("", leaf)


def format_category_cell(category_code: str, leaf_category: str) -> str:
    """
    Return: OCTOPIA | Catégorie | category_code | leaf_category
    """
    return f"OCTOPIA | Catégorie | {category_code} | {leaf_category}"


# ─────────────────────────────────────────
# TEMPLATE WRITER
# ─────────────────────────────────────────

def write_product_to_template(
    template_path: str,
    output_path: str,
    products: list[dict],
) -> str:
    """
    Load the Octopia xlsm template, append product rows, save to output_path.

    Each product dict must contain:
        title           str  – improved product title
        description     str  – raw scraped description (plain text)
        html_description str – LLM HTML description
        bullet_points   list – 5 marketing bullet points (stored in notes column)
        image_url       str  – primary image URL
        extra_images    list – up to 5 additional image URLs
        brand           str  – brand name (optional)
        gtin            str  – GTIN/EAN (optional)
        seller_ref      str  – seller reference (optional)
        category_name   str  – full category string from categories.csv
        category_code   str  – category code (optional, can be parsed from category_name)

    Returns the output_path on success.
    """
    # Copy template so we never mutate the original
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path, keep_vba=True)
    ws = wb.active  # Main product sheet

    # ── Find first empty data row (rows 1–10 are headers/meta) ──
    data_start_row = 11  # conservative default
    for r in range(9, ws.max_row + 2):
        if ws.cell(row=r, column=3).value is None:
            data_start_row = r
            break

    for idx, product in enumerate(products):
        row = data_start_row + idx

        # ── Category ───────────────────────────────────────────
        category_code = product.get("category_code", "")
        category_name = product.get("category_name", "")

        if not category_code:
            category_code, leaf = parse_category(category_name)
        else:
            _, leaf = parse_category(category_name)
            if not leaf:
                leaf = category_name.split("/")[-1].strip() if "/" in category_name else category_name

        # Update row 1 with category for this batch (last product wins if multiple)
        ws.cell(row=1, column=1).value = "OCTOPIA"
        ws.cell(row=1, column=2).value = "Catégorie"
        ws.cell(row=1, column=3).value = category_code
        ws.cell(row=1, column=4).value = leaf

        # ── Images ────────────────────────────────────────────
        extra = product.get("extra_images", [])
        image_cols = [5, 10, 11, 12, 13, 14]  # image 1–6
        all_images = [product.get("image_url", "")] + list(extra)

        # ── Write columns ─────────────────────────────────────
        ws.cell(row=row, column=1).value = product.get("gtin", "")
        ws.cell(row=row, column=2).value = product.get("seller_ref", "")
        ws.cell(row=row, column=3).value = product.get("title", "")[:132]
        ws.cell(row=row, column=4).value = product.get("description", "")[:2000]   # raw text
        ws.cell(row=row, column=8).value = product.get("brand", "")

        # Description marketing = HTML only
        html_desc = product.get("html_description", "")
        ws.cell(row=row, column=9).value = html_desc[:5000] if html_desc else ""

        # Images
        for i, col in enumerate(image_cols):
            if i < len(all_images) and all_images[i]:
                ws.cell(row=row, column=col).value = all_images[i]

        print(f"[data_mapper] Row {row} written: {product.get('title', '')[:60]}")

    wb.save(output_path)
    print(f"[data_mapper] Saved to: {output_path}")
    return output_path


# ─────────────────────────────────────────
# PIPELINE HELPER
# ─────────────────────────────────────────

def map_pipeline_result_to_template(
    template_path: str,
    scraped: dict,
    enhanced: dict,
    category_info: dict,
    output_path: str = None,
) -> str:
    """
    Convenience function: given outputs of the full pipeline, write one
    product row into the template and return the output file path.

    Args:
        template_path  – path to the base .xlsm template
        scraped        – dict returned by scraper.get_product_info()
        enhanced       – dict returned by openai_client.improve_product_content()
        category_info  – dict returned by category_utils.assign_category()
                         must include 'category_id' and 'category_name'
        output_path    – where to save; auto-generated if None

    Returns:
        output_path
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(template_path)[0]
        output_path = f"{base}_output_{ts}.xlsm"

    product = {
        "title":            enhanced.get("title", scraped.get("title", "")),
        "description":      scraped.get("description", ""),          # raw text → Description*
        "html_description": enhanced.get("html_description", ""),    # HTML → Description marketing
        "bullet_points":    enhanced.get("bullet_points", []),
        "image_url":        scraped.get("image_url", ""),
        "extra_images":     scraped.get("extra_images", []),
        "brand":            "",
        "gtin":             "",
        "seller_ref":       "",
        "category_name":    category_info.get("category_name", ""),
        "category_code":    str(category_info.get("category_id", "")),
    }

    return write_product_to_template(template_path, output_path, [product])
