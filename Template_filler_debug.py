from openpyxl import load_workbook
from datetime import datetime
import json


def debug_template_headers(template_path):
    """
    Find and display all headers in the template
    """
    try:
        wb = load_workbook(template_path)
        ws = wb.active
        
        print(f"\n{'='*80}")
        print(f"📋 Template: {ws.title}")
        print(f"{'='*80}\n")
        
        # Check rows 1-10 for headers
        for row_num in range(1, 11):
            print(f"\n🔍 ROW {row_num}:")
            print("-" * 80)
            
            headers = {}
            has_content = False
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col).value
                if cell_value:
                    has_content = True
                    headers[col] = cell_value
                    print(f"  Col {col}: {str(cell_value)[:60]}")
            
            if not has_content:
                print("  (Empty row)")
            
            # If this looks like header row, show count
            if has_content and len(headers) > 10:
                print(f"\n  ✅ ROW {row_num} looks like HEADER ROW ({len(headers)} columns)")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")


class TemplateFiller:
    """Handles filling and saving Excel templates with product data"""
    
    def __init__(self, template_path):
        """
        Initialize with template file
        
        Args:
            template_path: Path to the XLSM template file
        """
        self.template_path = template_path
        self.wb = None
        self.ws = None
    
    def load_template(self):
        """Load the Excel template"""
        try:
            self.wb = load_workbook(self.template_path)
            # Get first sheet (assumes category-specific sheet)
            self.ws = self.wb.active
            print(f"[template] ✅ Loaded template: {self.ws.title}")
            return True
        except Exception as e:
            print(f"[template] ERROR loading template: {e}")
            return False
    
    def find_header_row(self):
        """
        Find which row contains the headers
        Returns the row number or None if not found
        """
        print("\n[template] Searching for header row...")
        
        # Check rows 1-20
        for row_num in range(1, 21):
            headers_count = 0
            
            for col in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row=row_num, column=col).value
                if cell_value and isinstance(cell_value, str) and len(cell_value) > 3:
                    headers_count += 1
            
            # If row has many headers (>20 columns with text), it's probably the header row
            if headers_count > 20:
                print(f"[template] ✅ Found header row at ROW {row_num} ({headers_count} headers)")
                return row_num
        
        print("[template] ⚠️ Could not find header row automatically")
        return None
    
    def get_header_row(self, header_row_num=None):
        """
        Get the header row with column names
        
        Args:
            header_row_num: If None, tries to find it automatically
        """
        
        if header_row_num is None:
            header_row_num = self.find_header_row()
            if header_row_num is None:
                print("[template] ⚠️ Using default row 4")
                header_row_num = 4
        
        headers = {}
        
        for col in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(row=header_row_num, column=col).value
            if cell_value:
                headers[cell_value] = col
        
        print(f"[template] ℹ️ Found {len(headers)} headers in row {header_row_num}")
        return headers, header_row_num
    
    def fill_product_data(self, mapped_data, data_row_num=11):
        """
        Fill product data into the template
        
        Args:
            mapped_data: Dict with template fields as keys
            data_row_num: Row number to insert data (default: 11)
        """
        
        try:
            headers, header_row = self.get_header_row()
            
            print(f"\n[template] Filling {len(mapped_data)} fields into row {data_row_num}")
            filled_count = 0
            missing_count = 0
            
            for field_name, value in mapped_data.items():
                if field_name in headers:
                    col = headers[field_name]
                    cell = self.ws.cell(row=data_row_num, column=col)
                    
                    # Handle different data types
                    if isinstance(value, list):
                        cell.value = " | ".join(str(v) for v in value)
                    elif isinstance(value, dict):
                        cell.value = json.dumps(value, ensure_ascii=False)
                    else:
                        cell.value = value
                    
                    filled_count += 1
                else:
                    missing_count += 1
                    if missing_count <= 5:  # Show first 5 missing fields
                        print(f"  ⚠️ Header not found: {field_name}")
            
            print(f"[template] ✅ Filled {filled_count} fields, {missing_count} fields not found in template")
            return True
            
        except Exception as e:
            print(f"[template] ERROR filling data: {e}")
            return False
    
    def save_template(self, output_path):
        """
        Save the filled template to a new file
        
        Args:
            output_path: Path where to save the filled template
        """
        
        try:
            self.wb.save(output_path)
            print(f"[template] ✅ Saved filled template: {output_path}")
            return True
        except Exception as e:
            print(f"[template] ERROR saving template: {e}")
            return False
    
    def close(self):
        """Close the workbook"""
        if self.wb:
            self.wb.close()


def fill_template_for_product(template_path, mapped_data, product_id, output_dir="./filled_templates"):
    """
    Main function to fill a template with product data
    
    Args:
        template_path: Path to the XLSM template
        mapped_data: Dict with mapped product attributes
        product_id: Product ID for naming
        output_dir: Directory to save filled templates
        
    Returns:
        Path to saved template or None on error
    """
    
    import os
    
    # Create output directory if needed
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Initialize filler
    filler = TemplateFiller(template_path)
    
    if not filler.load_template():
        return None
    
    # Find header row automatically
    headers, header_row = filler.get_header_row()
    
    # Fill data (uses row 11 by default)
    if not filler.fill_product_data(mapped_data, data_row_num=11):
        filler.close()
        return None
    
    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"product_{product_id}_{timestamp}.xlsm"
    output_path = os.path.join(output_dir, output_filename)
    
    # Save template
    if not filler.save_template(output_path):
        filler.close()
        return None
    
    filler.close()
    return output_path


# ============================================================
# DEBUG SCRIPT
# ============================================================

if __name__ == "__main__":
    template_path = "pdt_template_fr-FR_20260305_090255.xlsm"
    
    if os.path.exists(template_path):
        print("🔍 DEBUGGING TEMPLATE HEADERS\n")
        debug_template_headers(template_path)
    else:
        print(f"❌ Template file not found: {template_path}")
